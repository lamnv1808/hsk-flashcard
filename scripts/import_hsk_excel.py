#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Import HSK vocabulary from source_data/HSK1-HSK6.xlsx into the static
hsk_flashcard_app/data.js used by the browser.

Design goals (see project brief):
  * Auto-detect every worksheet whose name starts with "HSK" (HSK7+ needs no change).
  * Required columns: B Chinese, C Pinyin, D Vietnamese(meaning),
    E Example, F Example Pinyin, G Example Translation.
  * Ignore blank rows; preserve UTF-8 / Chinese / Vietnamese / tone marks.
  * Emit the exact card schema already used:
    {id, level, word, pinyin, meaning, example, examplePinyin, translation}
  * PRESERVE all existing card IDs (anchored to the current data.js via the
    unique (level, word) key) so cloud/local progress never breaks.
  * Assign DETERMINISTIC new IDs starting at max(existing)+1, in
    numeric-level then row order.
  * Fail safely: validate everything in memory, only overwrite data.js
    (atomically) after validation succeeds. Print a summary.

Usage:
    python scripts/import_hsk_excel.py
"""

import json
import os
import re
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("ERROR: openpyxl is required. Install with:  python -m pip install openpyxl")

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SOURCE_XLSX = os.path.join(REPO_ROOT, "source_data", "HSK1-HSK6.xlsx")
OUT_JS = os.path.join(REPO_ROOT, "hsk_flashcard_app", "data.js")

# Card field order must match the existing schema exactly.
FIELDS = ["id", "level", "word", "pinyin", "meaning", "example",
          "examplePinyin", "translation"]

# Excel column indices (0-based) for the required columns B..G.
COL_WORD, COL_PINYIN, COL_MEANING = 1, 2, 3          # B, C, D
COL_EXAMPLE, COL_EXPINYIN, COL_TRANS = 4, 5, 6       # E, F, G
MIN_COLS = 7  # need columns A..G present


def cell(v):
    """Normalize a cell to a trimmed string (preserving all Unicode)."""
    if v is None:
        return ""
    return str(v).strip()


def level_num(name):
    """Numeric part of an HSK level name, for deterministic ordering."""
    m = re.search(r"\d+", name)
    return int(m.group()) if m else 0


def load_existing_ids(path):
    """Map (level, word) -> id from the current data.js, and the max id.
    Returns ({}, 0) if the file does not exist yet."""
    if not os.path.exists(path):
        return {}, 0
    text = open(path, encoding="utf-8").read()
    try:
        arr = json.loads(text[text.index("["):text.rindex("]") + 1])
    except ValueError:
        return {}, 0
    id_map, max_id = {}, 0
    for c in arr:
        key = (c.get("level"), c.get("word"))
        # Keep the first occurrence's id (keys are expected unique).
        id_map.setdefault(key, c["id"])
        if isinstance(c.get("id"), int):
            max_id = max(max_id, c["id"])
    return id_map, max_id


def read_sheet(ws, name):
    """Read a worksheet into a list of card dicts (without ids). Validates columns."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("sheet '%s' is empty" % name)
    # Validate that data rows have enough columns (A..G).
    width = max(len(r) for r in rows)
    if width < MIN_COLS:
        raise ValueError("sheet '%s' has only %d columns; need at least %d (A..G)"
                         % (name, width, MIN_COLS))
    cards = []
    for r in rows[1:]:  # skip header row
        r = list(r) + [None] * (MIN_COLS - len(r))
        word = cell(r[COL_WORD])
        if word == "":
            continue  # ignore blank rows (no Chinese word)
        cards.append({
            "level": name,
            "word": word,
            "pinyin": cell(r[COL_PINYIN]),
            "meaning": cell(r[COL_MEANING]),
            "example": cell(r[COL_EXAMPLE]),
            "examplePinyin": cell(r[COL_EXPINYIN]),
            "translation": cell(r[COL_TRANS]),
        })
    if not cards:
        raise ValueError("sheet '%s' has no data rows with a Chinese word" % name)
    return cards


def main():
    print("HSK importer")
    print("  source :", os.path.relpath(SOURCE_XLSX, REPO_ROOT))
    print("  output :", os.path.relpath(OUT_JS, REPO_ROOT))

    if not os.path.exists(SOURCE_XLSX):
        sys.exit("ERROR: source workbook not found: %s" % SOURCE_XLSX)

    wb = openpyxl.load_workbook(SOURCE_XLSX, data_only=True, read_only=True)
    hsk_sheets = [s for s in wb.sheetnames if s.strip().upper().startswith("HSK")]
    if not hsk_sheets:
        sys.exit("ERROR: no worksheet name starts with 'HSK'")
    # Deterministic order by numeric level (HSK1, HSK2, ... HSK10).
    hsk_sheets.sort(key=level_num)
    print("  sheets :", ", ".join(hsk_sheets))

    id_map, max_id = load_existing_ids(OUT_JS)
    print("  existing cards: %d (max id %d)" % (len(id_map), max_id))

    cards = []
    per_level = {}
    reused = 0
    next_id = max_id + 1
    try:
        for name in hsk_sheets:
            sheet_cards = read_sheet(wb[name], name)
            per_level[name] = len(sheet_cards)
            for c in sheet_cards:
                key = (c["level"], c["word"])
                if key in id_map:
                    cid = id_map[key]
                    reused += 1
                else:
                    cid = next_id
                    next_id += 1
                cards.append({
                    "id": cid, "level": c["level"], "word": c["word"],
                    "pinyin": c["pinyin"], "meaning": c["meaning"],
                    "example": c["example"], "examplePinyin": c["examplePinyin"],
                    "translation": c["translation"],
                })
    except ValueError as e:
        sys.exit("VALIDATION FAILED: %s (data.js NOT modified)" % e)

    # ---- Safety validations before writing ----
    ids = [c["id"] for c in cards]
    if len(set(ids)) != len(ids):
        sys.exit("VALIDATION FAILED: duplicate ids generated (data.js NOT modified)")
    # Every preserved key must keep its original id.
    for c in cards:
        key = (c["level"], c["word"])
        if key in id_map and id_map[key] != c["id"]:
            sys.exit("VALIDATION FAILED: id drift for %s (data.js NOT modified)" % (key,))
    # Sort by id so the array stays id-ordered (existing behavior).
    cards.sort(key=lambda c: c["id"])

    new_count = len(cards) - reused
    print("  per level     :", ", ".join("%s=%d" % (k, per_level[k]) for k in hsk_sheets))
    print("  total cards   :", len(cards))
    print("  reused ids    :", reused)
    print("  new ids       :", new_count, ("(%d..%d)" % (max_id + 1, next_id - 1)) if new_count else "")

    # Serialize in the same compact style as the current file.
    body = json.dumps([{f: c[f] for f in FIELDS} for c in cards],
                      ensure_ascii=False, separators=(",", ":"))
    header = ("// Generated from source_data/HSK1-HSK6.xlsx\n"
              "// DO NOT EDIT MANUALLY. Run: python scripts/import_hsk_excel.py\n")
    content = header + "window.HSK_CARDS = " + body + ";\n"

    # Atomic write: temp file then replace (only after all validation passed).
    tmp = OUT_JS + ".tmp"
    with open(tmp, "w", encoding="utf-8", newline="\n") as f:
        f.write(content)
    os.replace(tmp, OUT_JS)
    print("  wrote         :", os.path.relpath(OUT_JS, REPO_ROOT),
          "(%d bytes)" % os.path.getsize(OUT_JS))
    print("OK")


if __name__ == "__main__":
    main()
