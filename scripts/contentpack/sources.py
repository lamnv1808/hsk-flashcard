"""Source frontends: one .xlsx workbook (primary) or three CSV files (secondary).

Both frontends produce the SAME RawSource structure, which is what makes the
Canonical Source Intermediate identical for equivalent inputs.

Primary frontend  - a single .xlsx workbook with sheets:
                      manifest (required), cards (required), levels (optional)
Secondary frontend - a directory containing:
                      manifest.csv (required), cards.csv (required),
                      levels.csv (optional)
                    UTF-8, BOM tolerated, RFC 4180 parsing via the stdlib csv
                    module. No shell parsing, no spreadsheet application.

Fail-closed reading rules, each of which is a silent failure in the existing
HSK importer:
  - columns are bound by HEADER NAME, never by position
  - duplicate headers are fatal
  - unknown columns are fatal
  - hidden sheets and hidden rows are fatal
  - merged cells are fatal
  - formula cells are fatal (detected in a separate data_only=False pass;
    a cached formula result is an artifact of whoever last opened the file)
  - non-text cell values are fatal
"""

import csv
import hashlib
import io
import os

from . import schema
from .findings import Findings

MANIFEST_SECTION = "manifest"
CARDS_SECTION = "cards"
LEVELS_SECTION = "levels"

MANIFEST_HEADERS = ("key", "value")


class RawRow(object):
    """One source row, with per-cell coordinates preserved for reporting."""

    __slots__ = ("values", "coords", "source", "row_coord")

    def __init__(self, values, coords, source, row_coord):
        self.values = values        # column name -> raw str or None
        self.coords = coords        # column name -> coordinate string
        self.source = source        # sheet name or csv file name
        self.row_coord = row_coord  # "row 12" / "line 12"

    def coord(self, column):
        return self.coords.get(column, self.row_coord)


class RawSource(object):
    """Frontend-neutral parse result."""

    def __init__(self, kind):
        self.kind = kind            # "xlsx" | "csv"
        self.manifest = []          # list of (key, value, source, coord)
        self.cards = []             # list of RawRow
        self.levels = None          # list of RawRow, or None when absent
        self.files = []             # [{"name","bytes","sha256"}] for QA only


def _sha256_file(path):
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(65536), b""):
            h.update(chunk)
    return "sha256:" + h.hexdigest()


def _check_headers(raw_headers, allowed, required, section, findings, coord_of):
    """Bind columns by name. Duplicate or unknown headers are fatal."""
    seen = {}
    order = []
    for idx, name in enumerate(raw_headers):
        clean = (name or "").strip()
        if not clean:
            continue
        if clean in seen:
            findings.fatal(
                "DUPLICATE_HEADER",
                "duplicate column header '%s' in '%s'" % (clean, section),
                source=section, coord=coord_of(idx))
            continue
        if clean not in allowed:
            findings.fatal(
                "UNKNOWN_COLUMN",
                "unknown column '%s' in '%s'; allowed: %s"
                % (clean, section, ", ".join(sorted(allowed))),
                source=section, coord=coord_of(idx))
            continue
        seen[clean] = idx
        order.append(clean)

    for name in required:
        if name not in seen:
            findings.fatal(
                "MISSING_COLUMN",
                "required column '%s' is missing from '%s'" % (name, section),
                source=section)
    return seen, order


# --------------------------------------------------------------------------
# xlsx frontend
# --------------------------------------------------------------------------

def _col_letter(index_zero_based):
    """0 -> A, 25 -> Z, 26 -> AA."""
    n = index_zero_based + 1
    out = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        out = chr(ord("A") + rem) + out
    return out


def _load_openpyxl(findings):
    try:
        import openpyxl  # noqa: F401
        return openpyxl
    except ImportError:
        findings.fatal(
            "MISSING_DEPENDENCY",
            "openpyxl is required to read .xlsx sources. "
            "Install it with: python -m pip install openpyxl")
        return None


def _detect_formulas(openpyxl, path, sheet_names, findings):
    """Separate pass with data_only=False. Any formula cell is fatal.

    A cached formula value depends on whoever last recalculated the workbook,
    so trusting it would make the build depend on an editor's history rather
    than on the committed source.
    """
    wb = openpyxl.load_workbook(path, data_only=False, read_only=False)
    try:
        for name in sheet_names:
            if name not in wb.sheetnames:
                continue
            ws = wb[name]
            for row in ws.iter_rows():
                for cell in row:
                    value = cell.value
                    is_formula = (getattr(cell, "data_type", None) == "f") or (
                        isinstance(value, str) and value.startswith("="))
                    if is_formula:
                        findings.fatal(
                            "FORMULA_CELL",
                            "formula cells are not trusted as content; "
                            "replace it with a literal value",
                            source=name, coord=cell.coordinate)
    finally:
        wb.close()


def _sheet_rows(ws, section, findings):
    """Return non-empty data rows, rejecting hidden rows and merged cells."""
    if getattr(ws, "merged_cells", None) is not None:
        ranges = list(ws.merged_cells.ranges)
        if ranges:
            for rng in sorted(str(r) for r in ranges):
                findings.fatal(
                    "MERGED_CELL",
                    "merged cells silently drop data; unmerge range %s" % rng,
                    source=section, coord=rng)

    hidden = sorted(
        idx for idx, dim in ws.row_dimensions.items() if getattr(dim, "hidden", False))
    for idx in hidden:
        findings.fatal(
            "HIDDEN_ROW",
            "hidden rows are not imported silently; unhide or delete the row",
            source=section, coord="row %d" % idx)

    return list(ws.iter_rows())


def read_xlsx(path, findings):
    """Parse the primary .xlsx frontend."""
    src = RawSource("xlsx")

    size = os.path.getsize(path)
    if size > schema.MAX_SOURCE_BYTES:
        findings.fatal(
            "SOURCE_TOO_LARGE",
            "source is %d bytes, above the %d byte cap"
            % (size, schema.MAX_SOURCE_BYTES))
        return src
    src.files.append({
        "name": os.path.basename(path),
        "bytes": size,
        "sha256": _sha256_file(path),
    })

    openpyxl = _load_openpyxl(findings)
    if openpyxl is None:
        return src

    wanted = [MANIFEST_SECTION, CARDS_SECTION, LEVELS_SECTION]
    _detect_formulas(openpyxl, path, wanted, findings)
    if findings.has_fatal():
        return src

    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    try:
        for required in (MANIFEST_SECTION, CARDS_SECTION):
            if required not in wb.sheetnames:
                findings.fatal(
                    "MISSING_SHEET",
                    "required sheet '%s' is missing" % required)
        if findings.has_fatal():
            return src

        for name in wanted:
            if name in wb.sheetnames and wb[name].sheet_state != "visible":
                findings.fatal(
                    "HIDDEN_SHEET",
                    "sheet '%s' is %s; hidden sheets are not imported"
                    % (name, wb[name].sheet_state),
                    source=name)
        if findings.has_fatal():
            return src

        _read_xlsx_manifest(wb[MANIFEST_SECTION], src, findings)
        src.cards = _read_xlsx_table(
            wb[CARDS_SECTION], CARDS_SECTION,
            schema.ALL_CARD_COLUMNS,
            (schema.SOURCE_KEY_COLUMN,) + schema.REQUIRED_CARD_ROLES,
            findings)
        if LEVELS_SECTION in wb.sheetnames:
            src.levels = _read_xlsx_table(
                wb[LEVELS_SECTION], LEVELS_SECTION,
                schema.ALL_LEVEL_COLUMNS,
                schema.LEVELS_REQUIRED_COLUMNS,
                findings)
    finally:
        wb.close()

    return src


def _cell_text(cell, section, findings):
    """Cells must hold text. Numeric/date/bool values are fatal."""
    value = cell.value
    if value is None:
        return None
    if isinstance(value, str):
        return value
    findings.fatal(
        "UNSUPPORTED_CELL_TYPE",
        "cell holds %s; content cells must be text (format the column as Text)"
        % type(value).__name__,
        source=section, coord=cell.coordinate)
    return None


def _read_xlsx_manifest(ws, src, findings):
    rows = _sheet_rows(ws, MANIFEST_SECTION, findings)
    if findings.has_fatal():
        return
    if not rows:
        findings.fatal("EMPTY_SHEET", "sheet 'manifest' is empty",
                       source=MANIFEST_SECTION)
        return

    header = [_cell_text(c, MANIFEST_SECTION, findings) for c in rows[0]]
    bound, _ = _check_headers(
        header, set(MANIFEST_HEADERS), MANIFEST_HEADERS, MANIFEST_SECTION,
        findings, lambda i: "%s1" % _col_letter(i))
    if findings.has_fatal():
        return

    ki, vi = bound["key"], bound["value"]
    for row in rows[1:]:
        if ki >= len(row):
            continue
        key = _cell_text(row[ki], MANIFEST_SECTION, findings)
        value = _cell_text(row[vi], MANIFEST_SECTION, findings) if vi < len(row) else None
        if key is None or not key.strip():
            continue
        src.manifest.append(
            (key.strip(), value, MANIFEST_SECTION, row[ki].coordinate))


def _read_xlsx_table(ws, section, allowed, required, findings):
    rows = _sheet_rows(ws, section, findings)
    if findings.has_fatal():
        return []
    if not rows:
        findings.fatal("EMPTY_SHEET", "sheet '%s' is empty" % section, source=section)
        return []

    header = [_cell_text(c, section, findings) for c in rows[0]]
    bound, order = _check_headers(
        header, set(allowed), required, section, findings,
        lambda i: "%s1" % _col_letter(i))
    if findings.has_fatal():
        return []

    data = rows[1:]
    if len(data) > schema.MAX_SOURCE_ROWS:
        findings.fatal(
            "TOO_MANY_ROWS",
            "sheet '%s' has %d data rows, above the %d row cap"
            % (section, len(data), schema.MAX_SOURCE_ROWS),
            source=section)
        return []

    out = []
    for row in data:
        values, coords = {}, {}
        empty = True
        for name in order:
            idx = bound[name]
            cell = row[idx] if idx < len(row) else None
            text = _cell_text(cell, section, findings) if cell is not None else None
            values[name] = text
            coords[name] = cell.coordinate if cell is not None else section
            if text is not None and text.strip():
                empty = False
        if empty:
            continue
        row_no = row[0].row if row else 0
        out.append(RawRow(values, coords, section, "row %d" % row_no))
    return out


# --------------------------------------------------------------------------
# csv frontend
# --------------------------------------------------------------------------

def _read_csv_file(path, findings):
    """RFC 4180 parse. utf-8-sig tolerates an optional BOM."""
    size = os.path.getsize(path)
    if size > schema.MAX_SOURCE_BYTES:
        findings.fatal(
            "SOURCE_TOO_LARGE",
            "source is %d bytes, above the %d byte cap"
            % (size, schema.MAX_SOURCE_BYTES),
            source=os.path.basename(path))
        return None
    try:
        with io.open(path, "r", encoding="utf-8-sig", newline="") as fh:
            return list(csv.reader(fh))
    except UnicodeDecodeError as exc:
        findings.fatal(
            "MALFORMED_UNICODE",
            "file is not valid UTF-8: %s" % exc.reason,
            source=os.path.basename(path))
        return None
    except csv.Error as exc:
        findings.fatal(
            "MALFORMED_CSV",
            "CSV parse error: %s" % exc,
            source=os.path.basename(path))
        return None


def read_csv_dir(path, findings):
    """Parse the secondary CSV-directory frontend."""
    src = RawSource("csv")

    manifest_path = os.path.join(path, "manifest.csv")
    cards_path = os.path.join(path, "cards.csv")
    levels_path = os.path.join(path, "levels.csv")

    for required, name in ((manifest_path, "manifest.csv"), (cards_path, "cards.csv")):
        if not os.path.isfile(required):
            findings.fatal("MISSING_FILE", "required file '%s' is missing" % name)
    if findings.has_fatal():
        return src

    present = [manifest_path, cards_path]
    if os.path.isfile(levels_path):
        present.append(levels_path)
    for p in present:
        src.files.append({
            "name": os.path.basename(p),
            "bytes": os.path.getsize(p),
            "sha256": _sha256_file(p),
        })

    rows = _read_csv_file(manifest_path, findings)
    if rows is not None:
        _read_csv_manifest(rows, src, findings)

    rows = _read_csv_file(cards_path, findings)
    if rows is not None:
        src.cards = _read_csv_table(
            rows, "cards.csv", schema.ALL_CARD_COLUMNS,
            (schema.SOURCE_KEY_COLUMN,) + schema.REQUIRED_CARD_ROLES, findings)

    if os.path.isfile(levels_path):
        rows = _read_csv_file(levels_path, findings)
        if rows is not None:
            src.levels = _read_csv_table(
                rows, "levels.csv", schema.ALL_LEVEL_COLUMNS,
                schema.LEVELS_REQUIRED_COLUMNS, findings)

    return src


def _read_csv_manifest(rows, src, findings):
    if not rows:
        findings.fatal("EMPTY_FILE", "manifest.csv is empty", source="manifest.csv")
        return
    bound, _ = _check_headers(
        rows[0], set(MANIFEST_HEADERS), MANIFEST_HEADERS, "manifest.csv",
        findings, lambda i: "line 1 col %d" % (i + 1))
    if findings.has_fatal():
        return

    ki, vi = bound["key"], bound["value"]
    for line_no, row in enumerate(rows[1:], start=2):
        if ki >= len(row):
            continue
        key = row[ki]
        value = row[vi] if vi < len(row) else None
        if not key or not key.strip():
            continue
        src.manifest.append(
            (key.strip(), value, "manifest.csv", "line %d col %d" % (line_no, ki + 1)))


def _read_csv_table(rows, name, allowed, required, findings):
    if not rows:
        findings.fatal("EMPTY_FILE", "%s is empty" % name, source=name)
        return []
    bound, order = _check_headers(
        rows[0], set(allowed), required, name, findings,
        lambda i: "line 1 col %d" % (i + 1))
    if findings.has_fatal():
        return []

    data = rows[1:]
    if len(data) > schema.MAX_SOURCE_ROWS:
        findings.fatal(
            "TOO_MANY_ROWS",
            "%s has %d data rows, above the %d row cap"
            % (name, len(data), schema.MAX_SOURCE_ROWS),
            source=name)
        return []

    out = []
    for line_no, row in enumerate(data, start=2):
        values, coords = {}, {}
        empty = True
        for col in order:
            idx = bound[col]
            text = row[idx] if idx < len(row) else None
            values[col] = text
            coords[col] = "line %d col %d" % (line_no, idx + 1)
            if text is not None and text.strip():
                empty = False
        if empty:
            continue
        out.append(RawRow(values, coords, name, "line %d" % line_no))
    return out


# --------------------------------------------------------------------------

def read_source(path, findings):
    """Dispatch on the source path: a .xlsx file or a directory of CSVs."""
    if not os.path.exists(path):
        findings.fatal("SOURCE_NOT_FOUND", "source path does not exist: %s"
                       % os.path.basename(path))
        return None
    if os.path.isdir(path):
        return read_csv_dir(path, findings)
    if path.lower().endswith(".xlsx"):
        return read_xlsx(path, findings)
    findings.fatal(
        "UNSUPPORTED_SOURCE",
        "source must be a .xlsx workbook or a directory containing "
        "manifest.csv / cards.csv / optional levels.csv")
    return None
