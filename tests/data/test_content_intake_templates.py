#!/usr/bin/env python3
"""Phase 24F: the IELTS/TOEIC content intake templates stay honest and inert.

The templates are scaffolding a human fills in. Two things can go wrong and both
are silent, so both are asserted here:

  1. DRIFT -- the templates stop matching scripts/contentpack/schema.py, so an
     operator fills in a shape the pipeline will reject. Headers and reserved id
     ranges are therefore derived FROM the schema module, never re-typed.
  2. FABRICATION -- a placeholder that looks like real content or real
     provenance gets committed. An invented publisher or licence is a false
     legal claim; an invented word is course content FlashEdu did not author.
     So the templates must contain zero data rows and zero provenance values.

They must also be impossible to build or promote by accident: `.template.csv`
is not a filename the pipeline ever reads, and every template declares
draft/internal/not-visible.
"""

import csv
import io
import os
import sys

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.insert(0, os.path.join(ROOT, "scripts"))
sys.path.insert(0, os.path.join(ROOT, "tests", "support"))

from contentpack import schema                    # noqa: E402
from datajs import emit                           # noqa: E402

TEMPLATES = os.path.join(ROOT, "docs", "content", "templates")
GUIDE = os.path.join(ROOT, "docs", "content", "PHASE_24F_CONTENT_INTAKE.md")
APP = os.path.join(ROOT, "hsk_flashcard_app")

COURSES = ("ielts", "toeic")

# Anything that would make a template look like a shipped, launchable course.
FORBIDDEN_STATUS = {"status": "draft",
                    "launch.visible": "false",
                    "launch.readiness": "internal"}


def read_rows(path):
    with io.open(path, encoding="utf-8", newline="") as fh:
        return list(csv.reader(fh))


def manifest_map(path):
    rows = read_rows(path)
    return {r[0]: (r[1] if len(r) > 1 else "") for r in rows[1:] if r}


def main():
    fails = []

    def check(name, cond):
        if not cond:
            fails.append(name)

    # ---- the kit exists where the docs say it does ------------------------
    check("intake guide exists", os.path.isfile(GUIDE))
    check("templates README exists",
          os.path.isfile(os.path.join(TEMPLATES, "README.md")))

    for course in COURSES:
        d = os.path.join(TEMPLATES, course)
        man_p = os.path.join(d, "manifest.template.csv")
        lvl_p = os.path.join(d, "levels.template.csv")
        crd_p = os.path.join(d, "cards.template.csv")
        for label, p in (("manifest", man_p), ("levels", lvl_p), ("cards", crd_p)):
            check("%s: %s template exists" % (course, label), os.path.isfile(p))
        if not all(os.path.isfile(p) for p in (man_p, lvl_p, crd_p)):
            continue

        # ---- headers are derived from the schema, not re-typed ------------
        cards = read_rows(crd_p)
        check("%s: cards header matches schema.ALL_CARD_COLUMNS" % course,
              tuple(cards[0]) == schema.ALL_CARD_COLUMNS)
        for role in schema.REQUIRED_CARD_ROLES + (schema.SOURCE_KEY_COLUMN,):
            check("%s: cards header carries required '%s'" % (course, role),
                  role in cards[0])

        levels = read_rows(lvl_p)
        check("%s: levels header matches schema.ALL_LEVEL_COLUMNS" % course,
              tuple(levels[0]) == schema.ALL_LEVEL_COLUMNS)

        # ---- no authored content whatsoever -------------------------------
        check("%s: cards template has ZERO data rows" % course,
              len([r for r in cards[1:] if any(c.strip() for c in r)]) == 0)
        check("%s: levels template has ZERO data rows" % course,
              len([r for r in levels[1:] if any(c.strip() for c in r)]) == 0)

        # ---- manifest: keys are all schema-allowed ------------------------
        man = manifest_map(man_p)
        unknown = [k for k in man if k not in schema.MANIFEST_ALLOWED]
        check("%s: manifest declares no unknown key (%s)" % (course, unknown),
              unknown == [])
        missing = [k for k in schema.MANIFEST_REQUIRED if k not in man]
        check("%s: manifest carries every required key (%s)" % (course, missing),
              missing == [])
        computed = [k for k in man if k in schema.MANIFEST_TOOL_COMPUTED]
        check("%s: manifest supplies no tool-computed key (%s)" % (course, computed),
              computed == [])

        # ---- frozen identity ---------------------------------------------
        check("%s: packId is the course id" % course, man.get("packId") == course)
        check("%s: courseId is the course id" % course, man.get("courseId") == course)
        check("%s: courseType is exam" % course, man.get("courseType") == "exam")

        lo, hi = schema.RESERVED_RANGES[course]
        check("%s: idRange.min is the reserved %d" % (course, lo),
              man.get("idRange.min") == str(lo))
        check("%s: idRange.max is the reserved %d" % (course, hi),
              man.get("idRange.max") == str(hi))

        # ---- cannot claim launch ------------------------------------------
        for key, want in FORBIDDEN_STATUS.items():
            check("%s: %s is %s" % (course, key, want), man.get(key) == want)

        # ---- provenance is EMPTY, never invented --------------------------
        for key in schema.PROVENANCE_REQUIRED:
            check("%s: provenance '%s' is present but empty" % (course, key),
                  key in man and man[key].strip() == "")

    # ---- the kit cannot be consumed by the pipeline -----------------------
    consumable = []
    for base, _dirs, files in os.walk(TEMPLATES):
        for name in files:
            if name in ("manifest.csv", "cards.csv", "levels.csv"):
                consumable.append(os.path.join(base, name))
    check("no template is named like a real source file", consumable == [])

    # ---- the kit lives outside the runtime and outside source_data --------
    real = os.path.realpath(TEMPLATES)
    check("templates are not inside hsk_flashcard_app",
          not real.startswith(os.path.realpath(APP) + os.sep))
    check("templates are not inside source_data",
          not real.startswith(os.path.realpath(os.path.join(ROOT, "source_data")) + os.sep))

    # ---- no template leaked into the runtime or the catalog ---------------
    catalog_path = os.path.join(APP, "packs", "catalog.js")
    if os.path.isfile(catalog_path):
        with io.open(catalog_path, encoding="utf-8") as fh:
            catalog_text = fh.read()
        for course in COURSES:
            check("catalog still offers no %s pack" % course,
                  ('"packId":"%s"' % course) not in catalog_text)
    packs_dir = os.path.join(APP, "packs")
    entries = sorted(os.listdir(packs_dir)) if os.path.isdir(packs_dir) else []
    check("no course was promoted into the runtime",
          entries == ["catalog.js", "hsk"])

    # ---- still no real source dataset ------------------------------------
    src = os.path.join(ROOT, "source_data")
    listing = sorted(os.listdir(src)) if os.path.isdir(src) else []
    check("source_data still holds only the HSK workbook",
          listing == ["HSK1-HSK6.xlsx"])

    check_workbooks(check)

    return emit("content_intake_templates", fails)


def check_workbooks(check):
    """The .xlsx intake workbooks must be the CSV templates in Excel form.

    The founder will fill a workbook, so the workbook -- not just the CSV -- has
    to survive the real reader. Everything below is compared against the CSV
    templates or the schema module rather than a hand-typed expectation, so the
    two formats cannot drift apart.
    """
    try:
        import openpyxl
    except ImportError:                       # matches the suite-wide policy
        check("openpyxl available for workbook checks (skipped)", True)
        return

    sys.path.insert(0, os.path.join(ROOT, "scripts"))
    from contentpack import sources
    from contentpack.findings import Findings

    for course in COURSES:
        d = os.path.join(TEMPLATES, course)
        wb_path = os.path.join(d, "%s-intake.template.xlsx" % course)
        check("%s: intake workbook exists" % course, os.path.isfile(wb_path))
        if not os.path.isfile(wb_path):
            continue

        # ---- structural: exactly three visible sheets, no hidden anything --
        wb = openpyxl.load_workbook(wb_path, data_only=False)
        check("%s: workbook has exactly the three sheets" % course,
              wb.sheetnames == ["manifest", "cards", "levels"])
        for name in wb.sheetnames:
            ws = wb[name]
            check("%s/%s: sheet is visible" % (course, name),
                  ws.sheet_state == "visible")
            check("%s/%s: no merged cells" % (course, name),
                  len(list(ws.merged_cells.ranges)) == 0)
            hidden_rows = [i for i, dim in ws.row_dimensions.items() if dim.hidden]
            hidden_cols = [k for k, dim in ws.column_dimensions.items() if dim.hidden]
            check("%s/%s: no hidden rows" % (course, name), hidden_rows == [])
            check("%s/%s: no hidden columns" % (course, name), hidden_cols == [])
            # Every populated cell is a Text-formatted string. Excel otherwise
            # coerces 1000000 to a number and eats leading zeros in sourceKeys.
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    check("%s/%s: cell %s is a string" % (course, name, cell.coordinate),
                          isinstance(cell.value, str))
                    check("%s/%s: cell %s is Text-formatted" % (course, name, cell.coordinate),
                          cell.number_format == "@")
                    check("%s/%s: cell %s holds no formula" % (course, name, cell.coordinate),
                          not (isinstance(cell.value, str) and cell.value.startswith("=")))
        check("%s: workbook declares no external links" % course,
              not getattr(wb, "_external_links", []))
        check("%s: workbook is not macro-enabled" % course,
              wb_path.endswith(".xlsx"))

        # ---- the REAL reader accepts it -----------------------------------
        findings = Findings()
        parsed = sources.read_xlsx(wb_path, findings)
        fatal = findings.has_fatal()
        check("%s: real read_xlsx reports no fatal" % course, fatal is False)
        if fatal:
            # A fatal IS the failure. The reader abandons parsing and leaves
            # sections as None, so everything below would raise rather than
            # report -- skip to the next course instead of crashing the suite.
            continue

        check("%s: workbook carries zero card rows" % course,
              len(parsed.cards or []) == 0)
        check("%s: workbook carries zero level rows" % course,
              len(parsed.levels or []) == 0)

        parsed_manifest = {k: v for (k, v, _s, _c) in (parsed.manifest or [])}

        # ---- semantic parity with the CSV templates ------------------------
        csv_manifest = manifest_map(os.path.join(d, "manifest.template.csv"))
        # Empty CSV cells arrive as None from the reader; normalise both sides.
        def norm(m):
            return {k: (v or "").strip() for k, v in m.items()}
        check("%s: XLSX manifest equals the CSV manifest" % course,
              norm(parsed_manifest) == norm(csv_manifest))

        csv_cards = read_rows(os.path.join(d, "cards.template.csv"))
        csv_levels = read_rows(os.path.join(d, "levels.template.csv"))
        xl_cards = [c.value for c in next(wb["cards"].iter_rows(max_row=1))]
        xl_levels = [c.value for c in next(wb["levels"].iter_rows(max_row=1))]
        check("%s: XLSX cards header equals the CSV header" % course,
              xl_cards == csv_cards[0])
        check("%s: XLSX levels header equals the CSV header" % course,
              xl_levels == csv_levels[0])
        check("%s: XLSX cards header matches schema" % course,
              tuple(xl_cards) == schema.ALL_CARD_COLUMNS)
        check("%s: XLSX levels header matches schema" % course,
              tuple(xl_levels) == schema.ALL_LEVEL_COLUMNS)

        # ---- safety state survives the round trip --------------------------
        check("%s: workbook packId is the course" % course,
              parsed_manifest.get("packId") == course)
        lo, hi = schema.RESERVED_RANGES[course]
        check("%s: workbook idRange.min is reserved %d" % (course, lo),
              parsed_manifest.get("idRange.min") == str(lo))
        check("%s: workbook idRange.max is reserved %d" % (course, hi),
              parsed_manifest.get("idRange.max") == str(hi))
        for key, want in FORBIDDEN_STATUS.items():
            check("%s: workbook %s is %s" % (course, key, want),
                  parsed_manifest.get(key) == want)
        for key in schema.PROVENANCE_REQUIRED:
            check("%s: workbook provenance '%s' is empty" % (course, key),
                  (parsed_manifest.get(key) or "").strip() == "")
        computed = [k for k in parsed_manifest if k in schema.MANIFEST_TOOL_COMPUTED]
        check("%s: workbook supplies no tool-computed key (%s)" % (course, computed),
              computed == [])

        # ---- the workbook cannot be consumed by accident -------------------
        real_wb = os.path.realpath(wb_path)
        check("%s: workbook is outside source_data" % course,
              not real_wb.startswith(
                  os.path.realpath(os.path.join(ROOT, "source_data")) + os.sep))
        check("%s: workbook is outside the runtime" % course,
              not real_wb.startswith(os.path.realpath(APP) + os.sep))
        check("%s: workbook filename is a template, not a source name" % course,
              os.path.basename(wb_path).endswith(".template.xlsx"))


if __name__ == "__main__":
    sys.exit(main())
