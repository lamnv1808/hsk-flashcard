#!/usr/bin/env python3
"""Phase 24D - source parsing, header binding, and Excel/CSV equivalence.

Covers both frontends against the same committed fixture, and proves that an
.xlsx workbook and the CSV directory it was built from produce byte-identical
artifacts. Every fail-closed reading rule that the legacy HSK importer handles
silently (positional columns, hidden rows, merged cells, formulas, non-text
cells) is asserted to be fatal here.
"""

import os
import shutil
import sys
import tempfile

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.insert(0, os.path.join(ROOT, "scripts"))
sys.path.insert(0, os.path.join(ROOT, "tests", "support"))
sys.path.insert(0, os.path.join(ROOT, "tests", "fixtures", "packs"))

import packlib                                    # noqa: E402
from datajs import emit                           # noqa: E402
from contentpack.pipeline import Options, build   # noqa: E402

# openpyxl is required. A missing dependency must FAIL this suite rather than
# silently reporting it as passed or skipped, which would hide the entire
# Phase 24D coverage surface.
try:
    import openpyxl  # noqa: F401
    OPENPYXL_ERROR = None
except ImportError as exc:  # pragma: no cover - environment failure path
    OPENPYXL_ERROR = str(exc)


def codes(result):
    return {f.code for f in result.findings}


def run(source, out_dir, **kw):
    kw.setdefault("init_ledger", True)
    return build(Options(pack_id="demo", source=source, output=out_dir, **kw), ROOT)


def fresh(tmp, name="src"):
    dest = os.path.join(tmp, name)
    if os.path.isdir(dest):
        shutil.rmtree(dest)
    return packlib.copy_csv_source("demo", dest)


def main():
    fails = []

    def check(name, cond):
        if not cond:
            fails.append(name)

    if OPENPYXL_ERROR is not None:
        return emit("pack_build_parse",
                    ["openpyxl is required for Phase 24D suites: %s. "
                     "Install it with: python -m pip install openpyxl"
                     % OPENPYXL_ERROR])

    tmp = tempfile.mkdtemp(prefix="cpparse_")
    try:
        # --- valid CSV ---------------------------------------------------
        src = fresh(tmp)
        csv_result = run(src, os.path.join(tmp, "out_csv"))
        check("valid csv builds", not csv_result.findings.has_fatal())
        check("valid csv card count", len(csv_result.pack.cards) == 7)
        check("levels authored honored", csv_result.pack.levels_authored is True)
        check("declared optional roles",
              csv_result.pack.declared_roles ==
              ["pronunciation", "exampleText", "exampleTranslation", "tags"])

        # --- valid xlsx, built from the same CSVs -------------------------
        xlsx = packlib.make_xlsx(src, os.path.join(tmp, "demo.xlsx"))
        ledger = os.path.join(src, "demo-id-ledger.json")
        xlsx_result = run(xlsx, os.path.join(tmp, "out_xlsx"),
                          ledger_path=ledger, init_ledger=False)
        check("valid xlsx builds", not xlsx_result.findings.has_fatal())

        # --- exact equivalence -------------------------------------------
        check("csv/xlsx sourceChecksum identical",
              csv_result.source_checksum == xlsx_result.source_checksum)
        check("csv/xlsx contentChecksum identical",
              csv_result.content_checksum == xlsx_result.content_checksum)
        left = {a.name: a.data for a in csv_result.data_artifacts}
        right = {a.name: a.data for a in xlsx_result.data_artifacts}
        check("csv/xlsx data artifacts byte-identical", left == right)

        # --- Unicode is preserved exactly ---------------------------------
        by_key = {c["sourceKey"]: c for c in csv_result.pack.cards}
        check("chinese preserved", by_key["d-001"]["primaryPrompt"] == "爱")
        check("vietnamese diacritics preserved",
              by_key["d-001"]["definition"] == "yêu, thương")
        check("quoted comma survived csv parsing",
              "," in by_key["d-001"]["definition"])
        check("kana preserved",
              by_key["d-002"]["primaryPrompt"] == "ありがとう")
        check("pinyin tone mark preserved",
              by_key["d-006"]["pronunciation"] == "dǎ")
        check("hangul preserved",
              by_key["d-003"]["primaryPrompt"] == "한글")
        check("ipa preserved",
              by_key["d-004"]["pronunciation"] ==
              "/ˈwɔː.tər/")
        check("full-width punctuation preserved",
              by_key["d-001"]["exampleText"].endswith("。"))

        # --- UTF-8 BOM is tolerated ---------------------------------------
        src_bom = fresh(tmp, "src_bom")
        cards_path = os.path.join(src_bom, "cards.csv")
        with open(cards_path, "rb") as fh:
            body = fh.read()
        with open(cards_path, "wb") as fh:
            fh.write(b"\xef\xbb\xbf" + body)
        bom_result = run(src_bom, os.path.join(tmp, "out_bom"))
        check("utf-8 BOM tolerated", not bom_result.findings.has_fatal())
        check("BOM does not change content",
              bom_result.content_checksum == csv_result.content_checksum)

        # --- CRLF line terminators change nothing --------------------------
        # git normalizes these committed fixtures to CRLF on Windows checkout,
        # so the checksum must not depend on the line terminator.
        src_crlf = fresh(tmp, "src_crlf")
        for name in ("manifest.csv", "cards.csv", "levels.csv"):
            path = os.path.join(src_crlf, name)
            with open(path, "rb") as fh:
                raw = fh.read()
            with open(path, "wb") as fh:
                fh.write(raw.replace(b"\r\n", b"\n").replace(b"\n", b"\r\n"))
        crlf_result = run(src_crlf, os.path.join(tmp, "out_crlf"))
        check("CRLF source parses", not crlf_result.findings.has_fatal())
        check("CRLF does not change sourceChecksum",
              crlf_result.source_checksum == csv_result.source_checksum)
        check("CRLF does not change contentChecksum",
              crlf_result.content_checksum == csv_result.content_checksum)

        # --- embedded newline in a card cell is fatal ---------------------
        src_nl = fresh(tmp, "src_nl")
        packlib.edit_cards(src_nl, lambda rows: rows[:1] + [
            [rows[1][0], rows[1][1], "line1\nline2"] + rows[1][3:]] + rows[2:])
        nl_result = run(src_nl, os.path.join(tmp, "out_nl"))
        check("quoted newline in a card is fatal",
              "CONTROL_CHARACTER" in codes(nl_result))

        # --- structural failures ------------------------------------------
        src_missing = fresh(tmp, "src_missing")
        os.remove(os.path.join(src_missing, "manifest.csv"))
        check("missing manifest.csv is fatal",
              "MISSING_FILE" in codes(run(src_missing, os.path.join(tmp, "o1"))))

        src_col = fresh(tmp, "src_col")
        packlib.edit_cards(src_col, lambda rows: [
            [c for c in rows[0] if c != "definition"]] + [
            r[:3] + r[4:] for r in rows[1:]])
        check("missing required column is fatal",
              "MISSING_COLUMN" in codes(run(src_col, os.path.join(tmp, "o2"))))

        src_dup = fresh(tmp, "src_dup")
        packlib.edit_cards(src_dup, lambda rows: [
            rows[0] + ["deck"]] + [r + [""] for r in rows[1:]])
        check("duplicate header is fatal",
              "DUPLICATE_HEADER" in codes(run(src_dup, os.path.join(tmp, "o3"))))

        src_unk = fresh(tmp, "src_unk")
        packlib.edit_cards(src_unk, lambda rows: [
            rows[0] + ["mystery"]] + [r + ["x"] for r in rows[1:]])
        check("unknown column is fatal",
              "UNKNOWN_COLUMN" in codes(run(src_unk, os.path.join(tmp, "o4"))))

        src_mk = fresh(tmp, "src_mk")
        packlib.edit_manifest(src_mk, {"licence": "CC-BY"})
        check("unknown manifest key is fatal",
              "UNKNOWN_MANIFEST_KEY" in codes(run(src_mk, os.path.join(tmp, "o5"))))

        src_tc = fresh(tmp, "src_tc")
        packlib.edit_manifest(src_tc, {"cardCount": "7"})
        check("author-supplied tool-computed field is fatal",
              "TOOL_COMPUTED_FIELD" in codes(run(src_tc, os.path.join(tmp, "o6"))))

        src_deck = fresh(tmp, "src_deck")
        packlib.edit_cards(src_deck, lambda rows: rows[:1] + [
            [rows[1][0], "L9"] + rows[1][2:]] + rows[2:])
        check("undeclared deck is fatal",
              "UNDECLARED_DECK" in codes(run(src_deck, os.path.join(tmp, "o7"))))

        # --- levels sheet is optional -------------------------------------
        src_nolevels = fresh(tmp, "src_nolevels")
        os.remove(os.path.join(src_nolevels, "levels.csv"))
        nolevels = run(src_nolevels, os.path.join(tmp, "o8"))
        check("levels file is optional", not nolevels.findings.has_fatal())
        check("decks derived when levels absent",
              [e["deckId"] for e in nolevels.pack.levels] == ["L1", "L2"])
        check("derived levels are not marked authored",
              nolevels.pack.levels_authored is False)

        # --- malformed CSV -------------------------------------------------
        src_bad = fresh(tmp, "src_bad")
        with open(os.path.join(src_bad, "cards.csv"), "w", encoding="utf-8") as fh:
            fh.write("sourceKey,deck\n\"unterminated,L1\n")
        bad = run(src_bad, os.path.join(tmp, "o9"))
        check("malformed csv fails closed", bad.findings.has_fatal())

        # --- row cap --------------------------------------------------------
        src_big = fresh(tmp, "src_big")

        def explode(rows):
            head, first = rows[0], rows[1]
            out = [head]
            for i in range(20001):
                row = list(first)
                row[0] = "k-%05d" % i
                out.append(row)
            return out

        packlib.edit_cards(src_big, explode)
        check("row cap is enforced",
              "TOO_MANY_ROWS" in codes(run(src_big, os.path.join(tmp, "o10"))))

        # --- xlsx-specific hazards ------------------------------------------
        src_x = fresh(tmp, "src_x")

        formula_path = os.path.join(tmp, "formula.xlsx")
        packlib.make_xlsx(src_x, formula_path)
        wb = openpyxl.load_workbook(formula_path)
        wb["cards"]["D2"] = "=CONCATENATE(\"a\",\"b\")"
        wb.save(formula_path)
        wb.close()
        check("formula cell is fatal",
              "FORMULA_CELL" in codes(run(formula_path, os.path.join(tmp, "o11"))))

        hidden_sheet = os.path.join(tmp, "hidden_sheet.xlsx")
        packlib.make_xlsx(src_x, hidden_sheet)
        wb = openpyxl.load_workbook(hidden_sheet)
        wb["cards"].sheet_state = "hidden"
        wb.save(hidden_sheet)
        wb.close()
        check("hidden sheet is fatal",
              "HIDDEN_SHEET" in codes(run(hidden_sheet, os.path.join(tmp, "o12"))))

        hidden_row = os.path.join(tmp, "hidden_row.xlsx")
        packlib.make_xlsx(src_x, hidden_row)
        wb = openpyxl.load_workbook(hidden_row)
        wb["cards"].row_dimensions[3].hidden = True
        wb.save(hidden_row)
        wb.close()
        check("hidden row is fatal",
              "HIDDEN_ROW" in codes(run(hidden_row, os.path.join(tmp, "o13"))))

        merged = os.path.join(tmp, "merged.xlsx")
        packlib.make_xlsx(src_x, merged)
        wb = openpyxl.load_workbook(merged)
        wb["cards"].merge_cells("F4:G4")
        wb.save(merged)
        wb.close()
        check("merged cell is fatal",
              "MERGED_CELL" in codes(run(merged, os.path.join(tmp, "o14"))))

        numeric = os.path.join(tmp, "numeric.xlsx")
        packlib.make_xlsx(src_x, numeric)
        wb = openpyxl.load_workbook(numeric)
        wb["cards"]["C2"] = 12345
        wb.save(numeric)
        wb.close()
        check("non-text cell is fatal",
              "UNSUPPORTED_CELL_TYPE" in codes(run(numeric, os.path.join(tmp, "o15"))))

        no_sheet = os.path.join(tmp, "no_sheet.xlsx")
        packlib.make_xlsx(src_x, no_sheet)
        wb = openpyxl.load_workbook(no_sheet)
        del wb["cards"]
        wb.save(no_sheet)
        wb.close()
        check("missing sheet is fatal",
              "MISSING_SHEET" in codes(run(no_sheet, os.path.join(tmp, "o16"))))

    finally:
        shutil.rmtree(tmp, ignore_errors=True)

    return emit("pack_build_parse", fails)


if __name__ == "__main__":
    sys.exit(main())
