"""Fixture helpers for the Phase 24D pipeline suites.

The canonical fixture sources are the committed CSV files under
tests/fixtures/packs/<name>/. Equivalent .xlsx workbooks are BUILT AT TEST TIME
from those same CSVs rather than committed as binaries, so:

  - the fixture stays reviewable in a diff (a .xlsx is an opaque ZIP)
  - the Excel/CSV equivalence test compares a genuinely constructed workbook
    against the text source it was built from, instead of comparing two
    artifacts that a human kept in sync by hand
"""

import csv
import io
import os
import shutil

# .../tests/fixtures/packs
FIXTURE_ROOT = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(FIXTURE_ROOT)))

SECTION_FILES = (("manifest", "manifest.csv"),
                 ("cards", "cards.csv"),
                 ("levels", "levels.csv"))


def fixture_dir(name):
    return os.path.join(FIXTURE_ROOT, name)


def read_rows(path):
    with io.open(path, "r", encoding="utf-8-sig", newline="") as fh:
        return list(csv.reader(fh))


def write_rows(path, rows):
    parent = os.path.dirname(path)
    if parent and not os.path.isdir(parent):
        os.makedirs(parent)
    with io.open(path, "w", encoding="utf-8", newline="") as fh:
        csv.writer(fh, lineterminator="\n").writerows(rows)


def copy_csv_source(name, dest):
    """Copy a committed CSV fixture into a writable temp directory."""
    src = fixture_dir(name)
    if not os.path.isdir(dest):
        os.makedirs(dest)
    for _, filename in SECTION_FILES:
        path = os.path.join(src, filename)
        if os.path.isfile(path):
            shutil.copyfile(path, os.path.join(dest, filename))
    return dest


def make_xlsx(csv_dir, xlsx_path):
    """Build a workbook equivalent to a CSV source directory.

    Every cell is written as text so the workbook exercises the same
    text-only cell contract the pipeline enforces.
    """
    import openpyxl

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for section, filename in SECTION_FILES:
        path = os.path.join(csv_dir, filename)
        if not os.path.isfile(path):
            continue
        ws = wb.create_sheet(section)
        for row in read_rows(path):
            ws.append([("" if cell is None else str(cell)) for cell in row])
        for row in ws.iter_rows():
            for cell in row:
                cell.data_type = "s"
    parent = os.path.dirname(xlsx_path)
    if parent and not os.path.isdir(parent):
        os.makedirs(parent)
    wb.save(xlsx_path)
    wb.close()
    return xlsx_path


def edit_cards(csv_dir, mutate):
    """Apply a mutation to cards.csv rows in place. mutate(rows) -> rows."""
    path = os.path.join(csv_dir, "cards.csv")
    rows = read_rows(path)
    write_rows(path, mutate(rows))
    return path


def edit_manifest(csv_dir, updates, remove=()):
    """Set or remove manifest keys in place."""
    path = os.path.join(csv_dir, "manifest.csv")
    rows = read_rows(path)
    header, body = rows[0], rows[1:]
    body = [r for r in body if not r or r[0] not in remove]
    seen = {r[0]: i for i, r in enumerate(body) if r}
    for key in sorted(updates):
        if key in seen:
            body[seen[key]] = [key, updates[key]]
        else:
            body.append([key, updates[key]])
    write_rows(path, [header] + body)
    return path


def snapshot_tree(root):
    """Map relative path -> bytes, for proving a run wrote nothing."""
    out = {}
    if not os.path.isdir(root):
        return out
    for base, _dirs, files in os.walk(root):
        for name in files:
            full = os.path.join(base, name)
            with open(full, "rb") as fh:
                out[os.path.relpath(full, root).replace("\\", "/")] = fh.read()
    return out
